'''
TODO
prjName ending with .
prjCity from prjCode
'''

from openpyxl import load_workbook
from random import randint
from sys import argv
import json
import logging
import os
import time
import traceback

import LIB_MHSEARCH as libmh
import LIB_LOGCFG as liblog

liblog.setup_logging()

# XLSX
# wb = load_workbook(filename='mmr_null.xlsx', read_only=True)
# ws = wb['Sheet1']

# All Projects
# for row in ws.iter_rows('A{}:A{}'.format(ws.min_row,ws.max_row))

# for row in ws.get_squared_range(min_col=1, min_row=2, max_col=1, max_row=19):
    # for cell in row:
        # projectList.append(cell.value)

# CSV
with open(os.path.join('data', 'mmr_noresult.csv'), 'r') as fsv:
    mcsvl = fsv.readlines()

projectList = [pinfo.strip().split(',') for pinfo in mcsvl[1:]]

searches = {}
thisSearch = libmh.MahaSearch()

try:
    for prj in projectList:
        prjName = prj[0]
        builderName = prj[2]
        for cntr in range(7):
            try:
                thisCompare = {
                    'properName': prjName,
                    'label': '_value_'
                }
                searches[prjName] = thisSearch.searchResults('', includeScores=True, includeHighest=True, comparator=thisCompare) if builderName else {'_nopromoter_': prjName}
                break
            except:
                searches[prjName] = {'_failed_': prjName}
                continue
            finally:
                logging.info('Attempting '+ prjName)
                cntr += 1
                time.sleep(randint(8,20))
except:
    with open(os.path.join('_logs_', 'idetail_'+thisSearch.initiated.isoformat(timespec='hours')), 'w', encoding='utf-8') as dfl:
        traceback.print_exc(file=dfl)
finally:
# Save Results
    fname = 'MH_SearchResult_00' if len(argv)<=1 else argv[1]
    with open(os.path.join('results', fname + '.json'), 'w') as mfile:
        json.dump(searches, mfile, sort_keys=True, indent=4)